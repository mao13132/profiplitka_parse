import json
import random

from openpyxl import Workbook
from openpyxl.styles import Font

from datetime import datetime


class SaveResult:
    def __init__(self, good_dict):

        self.colums_checker = {}

        self.good_dict = good_dict['result']

        self.colums_harakt = ['Страна Бренда', 'Производитель', 'Серия', 'Страна производства внутреннего блока',
                              'Страна производства наружного блока', 'Завод изготовитель', 'Категории товара',
                              'На площадь (м2)', 'Охлаждение (кВт)', 'Обогрев (кВт)', 'Уровень шума (дБ)', 'Инвертор',
                              'Цвет внутреннего блока', 'Рабочий диапазон температур (охлаждение)',
                              'Рабочий диапазон температур (обогрев)', 'Хладагент', 'Класс энергоэффективности',
                              'Расход воздуха (м3/час)', 'С функциями', 'С фильтрами', 'Режим работы',
                              'Электропитание (В)', 'Потребляемая мощность (кВт)', 'Диаметр трубопроводов',
                              'Длина трассы (м)', 'Максимальный перепад высот (м)',
                              'Габариты внутреннего блока ШхВхГ (мм)', 'Габариты внешнего блока ШхВхГ (мм)',
                              'Вес внутреннего блока', 'Вес внешнего блока', 'Вес общий (кг.)', 'Wi-Fi',
                              'Подключение электропитания', 'Марка компрессора', 'Ионизатор воздуха',
                              'Увлажнение воздуха', 'Режим приточной вентиляции', 'Напор воздуха',
                              'Внешнее статическое давление (Па)', '', 'Дренажный насос', 'Зимний комплект',
                              'Размеры декоративной панели (мм)', 'Вес декоративной панели',
                              'Пульт дистанционного управления', 'Мобильный', 'Управление', 'Габариты, мм',
                              'УФ-лампа для обеззараживания', 'Виды блоков', 'Управление вертикальными жалюзи',
                              'По типу', 'MPN', 'Сенсор движения', 'Тип прецизионного кондиционера',
                              'Охлаждение конденсатора', 'Расход воды, м3/час', 'Компрессор', 'Количество компрессоров',
                              'Серия Belluna', 'Температурный режим', 'Объем камеры (м3)', 'Диапазон температур (°C)',
                              'Рабочее давление при +0 гр. в камере (бар)',
                              'Рабочее давление при +5 гр. в камере (бар)', 'Тип оттайки', 'Ток (А)', 'Тип установки',
                              'Функция естественного охаждения', 'Система Econavi', 'Размеры декоративной панели',
                              'Встроенный Bluetooth динамик']
        # self.colums_harakt = good_dict['name_colums']

        self.colums = ['ID продукта', 'Код', 'Имя продукта', 'Цена', 'Старая цена', 'Комплект главные',
                       'Комплект дополнительные', 'Фото', 'Категория', 'Производитель', 'Тип товара',
                       'Ссылка на сторонний сайт', 'Алгоритм', 'Ед.Измерения', 'ID', 'PARENT_ID', 'Видимость',
                       'Видимость варианта', 'Статус товара', 'Количество', 'Гарантия', 'Документы', 'Видео',
                       'Артикул', 'Описание']

    @staticmethod
    def save_to_json(filename, good_data):
        filename = f'{filename}.json'

        try:
            with open(filename, 'w', encoding='utf-8') as file:
                json.dump(good_data, file, indent=4, ensure_ascii=False)
        except:
            return False

        return filename

    def create_title(self, ws):

        global_count = 0

        for count, col in enumerate(self.colums):
            ws.cell(row=2, column=global_count + 1).value = col
            ws.cell(row=2, column=global_count + 1).font = Font(bold=True)

            global_count += 1

        for count, col in enumerate(self.colums_harakt):
            ws.cell(row=2, column=global_count + 1).value = col
            ws.cell(row=2, column=global_count + 1).font = Font(bold=True)

            self.colums_checker[col] = global_count + 1

            global_count += 1

        return self.colums_checker

    # def insert_contry(self, dict_in):
    #
    #     for key, value in dict_in.items():
    #         print(key)

    def write_data(self, ws, count_def, post):

        ws.cell(row=count_def, column=1).value = ''
        ws.cell(row=count_def, column=2).value = ''
        ws.cell(row=count_def, column=3).value = post['name']

        try:
            price = int(post['price'])
        except:
            try:
                price = post['price']
            except:
                price = ''

        ws.cell(row=count_def, column=4).value = price
        ws.cell(row=count_def, column=5).value = ''
        ws.cell(row=count_def, column=6).value = ''
        ws.cell(row=count_def, column=7).value = ''

        try:
            image = post['image']
        except:
            image = ''

        ws.cell(row=count_def, column=8).value = image

        try:
            category = post['xarakt']['Категории товара']
        except:
            category = ''

        ws.cell(row=count_def, column=9).value = category
        ws.cell(row=count_def, column=10).value = post['xarakt']['Производитель']
        ws.cell(row=count_def, column=11).value = category
        ws.cell(row=count_def, column=12).value = post['link']
        ws.cell(row=count_def, column=13).value = ''
        ws.cell(row=count_def, column=14).value = 'шт'
        ws.cell(row=count_def, column=15).value = ''
        ws.cell(row=count_def, column=16).value = ''
        ws.cell(row=count_def, column=17).value = 1
        ws.cell(row=count_def, column=18).value = 1
        ws.cell(row=count_def, column=19).value = ''
        ws.cell(row=count_def, column=20).value = 9999
        ws.cell(row=count_def, column=21).value = post['garant']
        ws.cell(row=count_def, column=22).value = post['documents']
        ws.cell(row=count_def, column=23).value = post['video']
        ws.cell(row=count_def, column=24).value = post['artikle']
        ws.cell(row=count_def, column=25).value = post['text']

        # for count_com, comment in enumerate(post['xarakt']):

        count = 0
        start_count = 26

        for key, value in post['xarakt'].items():
            # for key, value in self.colums_checker.items():
            ws.cell(row=count_def, column=self.colums_checker[key]).value = value
            # ws.cell(row=count_def + count, column=start_count).value = comment['author_comment']

            count += 1
            start_count += 1

        return True

    def itter_rows(self, ws):
        count_def = 3
        for count_post, post in enumerate(self.good_dict):
            if post['link'] == '':
                continue
            try:
                write_data = self.write_data(ws, count_def, post)
            except Exception as es:
                print(f'SaveResult: Исключение {es}')

            count_def += 1

        return True

    def create_number_uncolums(self, ws):

        global_count = 0
        start_count = 21

        for col in range(len(self.colums_harakt) + 5):
            ws.cell(row=1, column=start_count + global_count).value = random.randint(1111, 9999)

            global_count += 1

        return self.colums_checker

    def one_sheet(self, ws):

        self.create_number_uncolums(ws)

        name_colums_dict = self.create_title(ws)

        response_itter = self.itter_rows(ws)

        return True

    def save_file(self, filename):

        wb = Workbook()

        ws = wb.active

        result = self.one_sheet(ws)

        filename = f'{filename}'

        wb.save(f'{filename}.xlsx')

        # self.save_to_json(filename, self.good_dict)

        # print(f'Сохранил \n{filename}.xlsx\n{filename}.json')
        print(f'Сохранил \n{filename}.xlsx')

        return filename
